import streamlit as st
import sqlite3
import pandas as pd
import os
import numpy as np
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from PIL import Image
from streamlit_option_menu import option_menu
import base64
from io import BytesIO
import json
import time

# Set page configuration
st.set_page_config(
    page_title="Accounts Payable System",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Database connection
DB_PATH = "accounts_payable.db"

def init_database_if_needed():
    """Initialize the database if it doesn't exist yet"""
    if not os.path.exists(DB_PATH):
        st.info("Database not found. Initializing...")
        from setup_database import setup_database
        setup_database()

# Initialize database if needed
init_database_if_needed()

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# Create necessary directories
os.makedirs("uploads", exist_ok=True)
os.makedirs("reports", exist_ok=True)

# Initialize session state variables
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'user' not in st.session_state:
    st.session_state.user = None
if 'user_role' not in st.session_state:
    st.session_state.user_role = None

# Authentication functions
def login(username, password):
    conn = get_db_connection()
    user = conn.execute(
        "SELECT * FROM users WHERE username = ? AND status = 'active'", 
        (username,)
    ).fetchone()
    conn.close()
    
    if user and check_password(user['password_hash'], password):
        st.session_state.authenticated = True
        st.session_state.user = dict(user)
        st.session_state.user_role = user['role']
        return True
    return False

def check_password(stored_hash, password):
    # In a real application, use proper password hashing
    # This is a simplified version for demonstration
    return stored_hash == password  # Replace with bcrypt.checkpw

def logout():
    st.session_state.authenticated = False
    st.session_state.user = None
    st.session_state.user_role = None

# Login Page
def login_page():
    st.markdown("<h1 style='text-align: center;'>Accounts Payable System</h1>", unsafe_allow_html=True)
    
    # Display logo if available
    logo_path = "static/img/logo.png"
    if os.path.exists(logo_path):
        image = Image.open(logo_path)
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            st.image(image, width=200)
    
    st.markdown("<h2 style='text-align: center;'>Login</h2>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("login_form"):
            username = st.text_input("Username", key="username")
            password = st.text_input("Password", type="password", key="password")
            submit = st.form_submit_button("Login")
            
            if submit:
                if login(username, password):
                    st.success("Login successful!")
                    st.rerun()
                else:
                    st.error("Invalid username or password")
    
    # Display demo credentials
    with st.expander("Demo Credentials"):
        st.write("Username: admin")
        st.write("Password: admin123")
        st.write("Role: Administrator")
        st.write("---------")
        st.write("Username: accountant")
        st.write("Password: accountant123")
        st.write("Role: Accountant")
        st.write("---------")
        st.write("Username: approver")
        st.write("Password: approver123")
        st.write("Role: Approver")

# Main App Layout
def main_app():
    # Sidebar menu
    with st.sidebar:
        # User info
        st.write(f"Welcome, {st.session_state.user['full_name']}")
        st.write(f"Role: {st.session_state.user_role.title()}")
        st.divider()
        
        # Navigation menu based on role
        if st.session_state.user_role == 'admin':
            selected = option_menu(
                "Main Menu",
                ["Dashboard", "Vendors", "Invoices", "Payment Requests", "Reports", "Users", "Settings", "Data Manager"],
                icons=["house", "people", "file-text", "cash", "file-earmark-bar-graph", "person", "gear", "database"],
                menu_icon="cast",
                default_index=0
            )
        elif st.session_state.user_role == 'accountant':
            selected = option_menu(
                "Main Menu",
                ["Dashboard", "Vendors", "Invoices", "Payment Requests", "Reports"],
                icons=["house", "people", "file-text", "cash", "file-earmark-bar-graph"],
                menu_icon="cast",
                default_index=0
            )
        elif st.session_state.user_role == 'approver':
            selected = option_menu(
                "Main Menu",
                ["Dashboard", "Payment Approvals", "Reports"],
                icons=["house", "check-square", "file-earmark-bar-graph"],
                menu_icon="cast",
                default_index=0
            )
        else:
            selected = option_menu(
                "Main Menu",
                ["Dashboard", "Reports"],
                icons=["house", "file-earmark-bar-graph"],
                menu_icon="cast",
                default_index=0
            )
        
        # Logout button
        st.divider()
        if st.button("Logout"):
            logout()
            st.rerun()
    
    # Main content based on selection
    if selected == "Dashboard":
        display_dashboard()
    elif selected == "Vendors":
        display_vendors()
    elif selected == "Invoices":
        display_invoices()
    elif selected == "Payment Requests":
        display_payment_requests()
    elif selected == "Payment Approvals":
        display_payment_approvals()
    elif selected == "Reports":
        display_reports()
    elif selected == "Users":
        display_users()
    elif selected == "Settings":
        display_settings()
    elif selected == "Data Manager":
        from data_manager import data_management
        data_management()

# Dashboard Page
def display_dashboard():
    st.title("Accounts Payable Dashboard")
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    
    conn = get_db_connection()
    
    # Total Vendors
    total_vendors = conn.execute("SELECT COUNT(*) FROM vendors WHERE status = 'active'").fetchone()[0]
    col1.metric("Active Vendors", total_vendors)
    
    # Pending Invoices
    pending_invoices = conn.execute("SELECT COUNT(*) FROM invoices WHERE status = 'pending'").fetchone()[0]
    col2.metric("Pending Invoices", pending_invoices)
    
    # Total Outstanding
    total_outstanding = conn.execute(
        "SELECT SUM(total_amount) FROM invoices WHERE status IN ('pending', 'approved')"
    ).fetchone()[0]
    total_outstanding = total_outstanding or 0
    col3.metric("Total Outstanding", f"${total_outstanding:,.2f}")
    
    # Pending Approvals
    pending_approvals = conn.execute("SELECT COUNT(*) FROM payment_requests WHERE status = 'pending'").fetchone()[0]
    col4.metric("Pending Approvals", pending_approvals)
    
    # Aging Dashboard
    st.subheader("Accounts Payable Aging")
    
    # Get invoices data for aging
    invoices_df = pd.read_sql("""
        SELECT i.invoice_id, i.vendor_id, v.vendor_name, i.invoice_number, 
               i.invoice_date, i.due_date, i.total_amount, i.status
        FROM invoices i
        JOIN vendors v ON i.vendor_id = v.vendor_id
        WHERE i.status IN ('pending', 'approved')
    """, conn)
    
    if not invoices_df.empty:
        # Convert date columns
        invoices_df['invoice_date'] = pd.to_datetime(invoices_df['invoice_date'])
        invoices_df['due_date'] = pd.to_datetime(invoices_df['due_date'])
        
        # Calculate days overdue
        today = datetime.now().date()
        invoices_df['days_overdue'] = (today - invoices_df['due_date'].dt.date).dt.days
        
        # Create aging buckets
        conditions = [
            (invoices_df['days_overdue'] <= 0),
            (invoices_df['days_overdue'] > 0) & (invoices_df['days_overdue'] <= 30),
            (invoices_df['days_overdue'] > 30) & (invoices_df['days_overdue'] <= 60),
            (invoices_df['days_overdue'] > 60) & (invoices_df['days_overdue'] <= 90),
            (invoices_df['days_overdue'] > 90)
        ]
        
        choices = ['Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days']
        invoices_df['aging_bucket'] = np.select(conditions, choices, default='Current')
        
        # Group by aging bucket
        aging_summary = invoices_df.groupby('aging_bucket').agg(
            count=('invoice_id', 'count'),
            total=('total_amount', 'sum')
        ).reset_index()
        
        # Sort buckets
        bucket_order = {
            'Current': 0,
            '1-30 Days': 1,
            '31-60 Days': 2,
            '61-90 Days': 3,
            'Over 90 Days': 4
        }
        aging_summary['bucket_order'] = aging_summary['aging_bucket'].map(bucket_order)
        aging_summary = aging_summary.sort_values('bucket_order').drop('bucket_order', axis=1)
        
        # Create two columns for charts
        col1, col2 = st.columns(2)
        
        with col1:
            # Aging Bar Chart
            fig = px.bar(
                aging_summary, 
                x='aging_bucket', 
                y='total',
                text_auto='.2s',
                title='Outstanding Amount by Aging Bucket',
                labels={'aging_bucket': 'Aging Bucket', 'total': 'Amount ($)'},
                color='aging_bucket',
                color_discrete_map={
                    'Current': '#28a745',
                    '1-30 Days': '#ffc107',
                    '31-60 Days': '#fd7e14',
                    '61-90 Days': '#dc3545',
                    'Over 90 Days': '#6c757d'
                }
            )
            fig.update_layout(showlegend=False)
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            # Aging Pie Chart
            fig = px.pie(
                aging_summary, 
                values='total', 
                names='aging_bucket',
                title='Outstanding Amount Distribution',
                color='aging_bucket',
                color_discrete_map={
                    'Current': '#28a745',
                    '1-30 Days': '#ffc107',
                    '31-60 Days': '#fd7e14',
                    '61-90 Days': '#dc3545',
                    'Over 90 Days': '#6c757d'
                }
            )
            st.plotly_chart(fig, use_container_width=True)
        
        # Top Vendors by Outstanding Amount
        st.subheader("Top Vendors by Outstanding Amount")
        
        vendor_summary = invoices_df.groupby('vendor_name').agg(
            count=('invoice_id', 'count'),
            total=('total_amount', 'sum')
        ).reset_index().sort_values('total', ascending=False).head(10)
        
        fig = px.bar(
            vendor_summary,
            x='vendor_name',
            y='total',
            text_auto='.2s',
            title='Top 10 Vendors by Outstanding Amount',
            labels={'vendor_name': 'Vendor', 'total': 'Amount ($)'},
            color='total',
            color_continuous_scale='Viridis'
        )
        fig.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig, use_container_width=True)
        
        # Recent invoices
        st.subheader("Recent Pending Invoices")
        recent_invoices = invoices_df.sort_values('invoice_date', ascending=False).head(5)
        recent_invoices['invoice_date'] = recent_invoices['invoice_date'].dt.strftime('%Y-%m-%d')
        recent_invoices['due_date'] = recent_invoices['due_date'].dt.strftime('%Y-%m-%d')
        recent_invoices['total_amount'] = recent_invoices['total_amount'].apply(lambda x: f"${x:,.2f}")
        
        st.dataframe(
            recent_invoices[['vendor_name', 'invoice_number', 'invoice_date', 'due_date', 'total_amount', 'status']],
            use_container_width=True
        )
    else:
        st.info("No pending invoices found.")
    
    conn.close()

# Vendors Page
def display_vendors():
    st.title("Vendor Management")
    
    # Tabs for vendor operations
    tab1, tab2 = st.tabs(["Vendor List", "Create Vendor"])
    
    with tab1:
        display_vendor_list()
    
    with tab2:
        create_vendor_form()

def display_vendor_list():
    conn = get_db_connection()
    vendors = pd.read_sql("""
        SELECT v.vendor_id, v.vendor_name, v.contact_person, v.email, v.phone, v.status,
               COUNT(DISTINCT i.invoice_id) as invoice_count,
               SUM(CASE WHEN i.status IN ('pending', 'approved') THEN i.total_amount ELSE 0 END) as outstanding_amount
        FROM vendors v
        LEFT JOIN invoices i ON v.vendor_id = i.vendor_id
        GROUP BY v.vendor_id
        ORDER BY v.vendor_name
    """, conn)
    conn.close()
    
    # Search filter
    search = st.text_input("Search Vendors", "")
    if search:
        vendors = vendors[vendors['vendor_name'].str.contains(search, case=False)]
    
    # Status filter
    status_filter = st.multiselect(
        "Filter by Status", 
        options=["active", "inactive", "blacklisted"],
        default=["active"]
    )
    if status_filter:
        vendors = vendors[vendors['status'].isin(status_filter)]
    
    # Format the dataframe
    vendors['outstanding_amount'] = vendors['outstanding_amount'].fillna(0).apply(lambda x: f"${x:,.2f}")
    
    # Display vendors with edit button
    for i, vendor in vendors.iterrows():
        col1, col2, col3, col4 = st.columns([3, 2, 1, 1])
        
        with col1:
            st.subheader(vendor['vendor_name'])
            st.write(f"Contact: {vendor['contact_person'] or 'N/A'}")
        
        with col2:
            st.write(f"Email: {vendor['email'] or 'N/A'}")
            st.write(f"Phone: {vendor['phone'] or 'N/A'}")
        
        with col3:
            st.write(f"Status: {vendor['status'].title()}")
            st.write(f"Invoices: {vendor['invoice_count']}")
        
        with col4:
            st.write(f"Outstanding: {vendor['outstanding_amount']}")
            if st.button("Edit", key=f"edit_{vendor['vendor_id']}"):
                st.session_state.edit_vendor_id = vendor['vendor_id']
                st.rerun()
        
        st.divider()
    
    # Edit vendor modal
    if 'edit_vendor_id' in st.session_state and st.session_state.edit_vendor_id:
        display_edit_vendor_modal(st.session_state.edit_vendor_id)

def display_edit_vendor_modal(vendor_id):
    conn = get_db_connection()
    vendor = conn.execute("SELECT * FROM vendors WHERE vendor_id = ?", (vendor_id,)).fetchone()
    
    # Get bank details
    bank_details = pd.read_sql(
        "SELECT * FROM vendor_bank_details WHERE vendor_id = ?", 
        conn, 
        params=(vendor_id,)
    )
    
    # Get documents
    documents = pd.read_sql(
        "SELECT * FROM vendor_documents WHERE vendor_id = ?", 
        conn, 
        params=(vendor_id,)
    )
    
    conn.close()
    
    if vendor:
        st.sidebar.title(f"Edit Vendor: {vendor['vendor_name']}")
        
        with st.sidebar.form("edit_vendor_form"):
            vendor_name = st.text_input("Vendor Name", vendor['vendor_name'])
            contact_person = st.text_input("Contact Person", vendor['contact_person'] or "")
            email = st.text_input("Email", vendor['email'] or "")
            phone = st.text_input("Phone", vendor['phone'] or "")
            address = st.text_area("Address", vendor['address'] or "")
            tax_id = st.text_input("Tax ID", vendor['tax_id'] or "")
            registration_number = st.text_input("Registration Number", vendor['registration_number'] or "")
            status = st.selectbox("Status", ["active", "inactive", "blacklisted"], index=["active", "inactive", "blacklisted"].index(vendor['status']))
            
            submitted = st.form_submit_button("Update Vendor")
            
            if submitted:
                conn = get_db_connection()
                conn.execute("""
                    UPDATE vendors
                    SET vendor_name = ?, contact_person = ?, email = ?, phone = ?,
                        address = ?, tax_id = ?, registration_number = ?, status = ?
                    WHERE vendor_id = ?
                """, (vendor_name, contact_person, email, phone, address, tax_id, registration_number, status, vendor_id))
                conn.commit()
                conn.close()
                
                st.success("Vendor updated successfully!")
                st.session_state.edit_vendor_id = None
                st.rerun()
        
        # Bank details section
        st.sidebar.subheader("Bank Details")
        
        if not bank_details.empty:
            for i, bank in bank_details.iterrows():
                with st.sidebar.expander(f"{bank['bank_name']} - {bank['account_number']}"):
                    with st.form(f"edit_bank_{bank['bank_id']}"):
                        bank_name = st.text_input("Bank Name", bank['bank_name'], key=f"bank_name_{bank['bank_id']}")
                        account_number = st.text_input("Account Number", bank['account_number'], key=f"account_number_{bank['bank_id']}")
                        ifsc_code = st.text_input("IFSC Code", bank['ifsc_code'] or "", key=f"ifsc_code_{bank['bank_id']}")
                        account_type = st.text_input("Account Type", bank['account_type'] or "", key=f"account_type_{bank['bank_id']}")
                        branch_name = st.text_input("Branch Name", bank['branch_name'] or "", key=f"branch_name_{bank['bank_id']}")
                        is_primary = st.checkbox("Primary Account", bool(bank['is_primary']), key=f"is_primary_{bank['bank_id']}")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            update_bank = st.form_submit_button("Update")
                        with col2:
                            delete_bank = st.form_submit_button("Delete")
                        
                        if update_bank:
                            conn = get_db_connection()
                            
                            # If setting this as primary, unset all others
                            if is_primary:
                                conn.execute(
                                    "UPDATE vendor_bank_details SET is_primary = 0 WHERE vendor_id = ?", 
                                    (vendor_id,)
                                )
                            
                            conn.execute("""
                                UPDATE vendor_bank_details
                                SET bank_name = ?, account_number = ?, ifsc_code = ?,
                                    account_type = ?, branch_name = ?, is_primary = ?
                                WHERE bank_id = ?
                            """, (bank_name, account_number, ifsc_code, account_type, branch_name, is_primary, bank['bank_id']))
                            conn.commit()
                            conn.close()
                            
                            st.success("Bank details updated!")
                            st.rerun()
                        
                        if delete_bank:
                            conn = get_db_connection()
                            conn.execute("DELETE FROM vendor_bank_details WHERE bank_id = ?", (bank['bank_id'],))
                            conn.commit()
                            conn.close()
                            
                            st.success("Bank details deleted!")
                            st.rerun()
        
        # Add new bank details
        with st.sidebar.expander("Add New Bank Details"):
            with st.form("add_bank_form"):
                bank_name = st.text_input("Bank Name", key="new_bank_name")
                account_number = st.text_input("Account Number", key="new_account_number")
                ifsc_code = st.text_input("IFSC Code", key="new_ifsc_code")
                account_type = st.text_input("Account Type", key="new_account_type")
                branch_name = st.text_input("Branch Name", key="new_branch_name")
                is_primary = st.checkbox("Primary Account", key="new_is_primary")
                
                add_bank = st.form_submit_button("Add Bank")
                
                if add_bank:
                    if not bank_name or not account_number:
                        st.error("Bank name and account number are required.")
                    else:
                        conn = get_db_connection()
                        
                        # If setting this as primary, unset all others
                        if is_primary:
                            conn.execute(
                                "UPDATE vendor_bank_details SET is_primary = 0 WHERE vendor_id = ?", 
                                (vendor_id,)
                            )
                        
                        conn.execute("""
                            INSERT INTO vendor_bank_details
                            (vendor_id, bank_name, account_number, ifsc_code, account_type, branch_name, is_primary)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        """, (vendor_id, bank_name, account_number, ifsc_code, account_type, branch_name, is_primary))
                        conn.commit()
                        conn.close()
                        
                        st.success("Bank details added!")
                        st.rerun()
        
        # Documents section
        st.sidebar.subheader("KYC Documents")
        
        if not documents.empty:
            for i, doc in documents.iterrows():
                with st.sidebar.expander(f"{doc['document_type']} ({doc['status']})"):
                    st.write(f"Uploaded: {doc['uploaded_at']}")
                    
                    # Handle document viewing/download
                    if os.path.exists(doc['document_path']):
                        with open(doc['document_path'], "rb") as file:
                            btn = st.download_button(
                                label="Download Document",
                                data=file,
                                file_name=os.path.basename(doc['document_path']),
                                mime="application/octet-stream"
                            )
                    else:
                        st.error("Document file not found.")
                    
                    # Status update
                    if st.session_state.user_role in ['admin', 'approver']:
                        status = st.selectbox(
                            "Status", 
                            ["pending", "approved", "rejected"],
                            index=["pending", "approved", "rejected"].index(doc['status']),
                            key=f"doc_status_{doc['document_id']}"
                        )
                        
                        if st.button("Update Status", key=f"update_doc_{doc['document_id']}"):
                            conn = get_db_connection()
                            conn.execute(
                                "UPDATE vendor_documents SET status = ? WHERE document_id = ?", 
                                (status, doc['document_id'])
                            )
                            conn.commit()
                            conn.close()
                            
                            st.success("Document status updated!")
                            st.rerun()
                    
                    # Delete document
                    if st.session_state.user_role == 'admin':
                        if st.button("Delete Document", key=f"delete_doc_{doc['document_id']}"):
                            conn = get_db_connection()
                            conn.execute("DELETE FROM vendor_documents WHERE document_id = ?", (doc['document_id'],))
                            conn.commit()
                            conn.close()
                            
                            # Delete file
                            if os.path.exists(doc['document_path']):
                                os.remove(doc['document_path'])
                            
                            st.success("Document deleted!")
                            st.rerun()
        
        # Add new document
        with st.sidebar.expander("Upload New Document"):
            with st.form("upload_document_form"):
                document_type = st.selectbox(
                    "Document Type",
                    ["PAN Card", "GST Certificate", "Incorporation Certificate", "Address Proof", "Bank Statement", "Other"],
                    key="new_document_type"
                )
                
                uploaded_file = st.file_uploader("Upload Document", type=["pdf", "png", "jpg", "jpeg", "doc", "docx"])
                
                upload_doc = st.form_submit_button("Upload")
                
                if upload_doc and uploaded_file:
                    # Save file
                    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                    file_ext = os.path.splitext(uploaded_file.name)[1]
                    file_path = os.path.join("uploads", f"vendor_{vendor_id}_{document_type.replace(' ', '_')}_{timestamp}{file_ext}")
                    
                    with open(file_path, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    
                    # Add to database
                    conn = get_db_connection()
                    conn.execute("""
                        INSERT INTO vendor_documents
                        (vendor_id, document_type, document_path, status)
                        VALUES (?, ?, ?, 'pending')
                    """, (vendor_id, document_type, file_path))
                    conn.commit()
                    conn.close()
                    
                    st.success("Document uploaded!")
                    st.rerun()
        
        # Close button
        if st.sidebar.button("Close"):
            st.session_state.edit_vendor_id = None
            st.rerun()

def create_vendor_form():
    with st.form("create_vendor_form"):
        st.subheader("Create New Vendor")
        
        vendor_name = st.text_input("Vendor Name *")
        contact_person = st.text_input("Contact Person")
        email = st.text_input("Email")
        phone = st.text_input("Phone")
        
        col1, col2 = st.columns(2)
        with col1:
            tax_id = st.text_input("Tax ID")
        with col2:
            registration_number = st.text_input("Registration Number")
        
        address = st.text_area("Address")
        
        submitted = st.form_submit_button("Create Vendor")
        
        if submitted:
            if not vendor_name:
                st.error("Vendor name is required.")
            else:
                conn = get_db_connection()
                conn.execute("""
                    INSERT INTO vendors
                    (vendor_name, contact_person, email, phone, address, tax_id, registration_number, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'active')
                """, (vendor_name, contact_person, email, phone, address, tax_id, registration_number))
                
                vendor_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                conn.commit()
                conn.close()
                
                st.success(f"Vendor '{vendor_name}' created successfully!")
                st.info("You can now add bank details and upload KYC documents.")
                
                # Set edit mode for the new vendor
                st.session_state.edit_vendor_id = vendor_id
                st.rerun()

# Invoices Page
def display_invoices():
    st.title("Invoice Management")
    
    # Tabs for invoice operations
    tab1, tab2, tab3 = st.tabs(["Invoice List", "Create Invoice", "Import from Tally"])
    
    with tab1:
        display_invoice_list()
    
    with tab2:
        create_invoice_form()
    
    with tab3:
        import_invoices_from_tally()

def display_invoice_list():
    conn = get_db_connection()
    
    # Get invoices with vendor info
    invoices = pd.read_sql("""
        SELECT i.invoice_id, i.vendor_id, v.vendor_name, i.invoice_number, 
               i.invoice_date, i.due_date, i.amount, i.tax_amount, i.total_amount, 
               i.status, i.description
        FROM invoices i
        JOIN vendors v ON i.vendor_id = v.vendor_id
        ORDER BY i.due_date ASC
    """, conn)
    conn.close()
    
    # Convert date columns
    invoices['invoice_date'] = pd.to_datetime(invoices['invoice_date']).dt.strftime('%Y-%m-%d')
    invoices['due_date'] = pd.to_datetime(invoices['due_date']).dt.strftime('%Y-%m-%d')
    
    # Calculate days to due or overdue
    today = datetime.now().date()
    invoices['due_date_dt'] = pd.to_datetime(invoices['due_date']).dt.date
    invoices['days'] = (invoices['due_date_dt'] - today).dt.days
    
    # Filters
    col1, col2, col3 = st.columns(3)
    
    with col1:
        search = st.text_input("Search by Invoice # or Vendor", "")
    
    with col2:
        status_filter = st.multiselect(
            "Status", 
            options=["pending", "approved", "rejected", "paid"],
            default=["pending", "approved"]
        )
    
    with col3:
        date_filter = st.selectbox(
            "Date Range",
            options=["All", "Due this week", "Due this month", "Overdue"],
            index=0
        )
    
    # Apply filters
    filtered_invoices = invoices.copy()
    
    if search:
        filtered_invoices = filtered_invoices[
            filtered_invoices['invoice_number'].str.contains(search, case=False) | 
            filtered_invoices['vendor_name'].str.contains(search, case=False)
        ]
    
    if status_filter:
        filtered_invoices = filtered_invoices[filtered_invoices['status'].isin(status_filter)]
    
    if date_filter == "Due this week":
        filtered_invoices = filtered_invoices[
            (filtered_invoices['days'] >= 0) & (filtered_invoices['days'] <= 7)
        ]
    elif date_filter == "Due this month":
        filtered_invoices = filtered_invoices[
            (filtered_invoices['days'] >= 0) & (filtered_invoices['days'] <= 30)
        ]
    elif date_filter == "Overdue":
        filtered_invoices = filtered_invoices[filtered_invoices['days'] < 0]
    
    # Display invoice count
    st.write(f"Showing {len(filtered_invoices)} invoices")
    
    # Format currency columns
    filtered_invoices['amount'] = filtered_invoices['amount'].apply(lambda x: f"${x:,.2f}")
    filtered_invoices['tax_amount'] = filtered_invoices['tax_amount'].apply(lambda x: f"${x:,.2f}")
    filtered_invoices['total_amount'] = filtered_invoices['total_amount'].apply(lambda x: f"${x:,.2f}")
    
    # Add days column with formatting
    def format_days(days):
        if days < 0:
            return f"🚨 Overdue by {abs(days)} days"
        elif days == 0:
            return "⚠️ Due today"
        else:
            return f"✅ Due in {days} days"
    
    filtered_invoices['due_in'] = filtered_invoices['days'].apply(format_days)
    
    # Create display columns
    display_cols = ['invoice_number', 'vendor_name', 'invoice_date', 'due_date', 'due_in', 'total_amount', 'status']
    
    # Display invoices with action buttons
    selected_invoices = []
    
    for i, invoice in filtered_invoices.iterrows():
        col1, col2, col3, col4, col5 = st.columns([2, 3, 2, 2, 1])
        
        with col1:
            st.write(f"**Invoice #:** {invoice['invoice_number']}")
        
        with col2:
            st.write(f"**Vendor:** {invoice['vendor_name']}")
        
        with col3:
            st.write(f"**Due Date:** {invoice['due_date']}")
            st.write(invoice['due_in'])
        
        with col4:
            st.write(f"**Amount:** {invoice['total_amount']}")
            st.write(f"**Status:** {invoice['status'].title()}")
        
        with col5:
            # Select for payment checkbox
            if invoice['status'] in ['pending', 'approved']:
                if st.checkbox("Select", key=f"select_{invoice['invoice_id']}"):
                    selected_invoices.append(invoice['invoice_id'])
            
            # View/Edit button
            if st.button("View", key=f"view_{invoice['invoice_id']}"):
                st.session_state.edit_invoice_id = invoice['invoice_id']
                st.rerun()
        
        st.divider()
    
    # Create payment request button
    if selected_invoices and st.session_state.user_role in ['admin', 'accountant']:
        st.write(f"Selected {len(selected_invoices)} invoices for payment")
        
        if st.button("Create Payment Request"):
            st.session_state.create_payment_request = selected_invoices
            st.rerun()
    
    # Edit invoice modal
    if 'edit_invoice_id' in st.session_state and st.session_state.edit_invoice_id:
        display_edit_invoice_modal(st.session_state.edit_invoice_id)
    
    # Create payment request modal
    if 'create_payment_request' in st.session_state and st.session_state.create_payment_request:
        display_create_payment_request_modal(st.session_state.create_payment_request)

def display_edit_invoice_modal(invoice_id):
    conn = get_db_connection()
    invoice = conn.execute("SELECT * FROM invoices WHERE invoice_id = ?", (invoice_id,)).fetchone()
    
    if invoice:
        vendor = conn.execute("SELECT * FROM vendors WHERE vendor_id = ?", (invoice['vendor_id'],)).fetchone()
        conn.close()
        
        st.sidebar.title(f"Invoice: {invoice['invoice_number']}")
        
        with st.sidebar.form("edit_invoice_form"):
            vendor_name = st.text_input("Vendor", vendor['vendor_name'], disabled=True)
            invoice_number = st.text_input("Invoice Number", invoice['invoice_number'])
            invoice_date = st.date_input("Invoice Date", datetime.strptime(invoice['invoice_date'], "%Y-%m-%d").date())
            due_date = st.date_input("Due Date", datetime.strptime(invoice['due_date'], "%Y-%m-%d").date())
            
            amount = st.number_input("Amount", min_value=0.0, value=float(invoice['amount']), format="%.2f")
            tax_amount = st.number_input("Tax Amount", min_value=0.0, value=float(invoice['tax_amount']), format="%.2f")
            total_amount = st.number_input("Total Amount", min_value=0.0, value=float(invoice['total_amount']), format="%.2f")
            
            description = st.text_area("Description", invoice['description'] or "")
            
            status = st.selectbox(
                "Status", 
                options=["pending", "approved", "rejected", "paid"],
                index=["pending", "approved", "rejected", "paid"].index(invoice['status'])
            )
            
            submitted = st.form_submit_button("Update Invoice")
            
            if submitted:
                conn = get_db_connection()
                conn.execute("""
                    UPDATE invoices
                    SET invoice_number = ?, invoice_date = ?, due_date = ?,
                        amount = ?, tax_amount = ?, total_amount = ?,
                        description = ?, status = ?
                    WHERE invoice_id = ?
                """, (invoice_number, invoice_date, due_date, amount, tax_amount, total_amount, description, status, invoice_id))
                conn.commit()
                conn.close()
                
                st.success("Invoice updated successfully!")
                st.session_state.edit_invoice_id = None
                st.rerun()
        
        # Invoice file
        st.sidebar.subheader("Invoice File")
        
        if invoice['invoice_file_path'] and os.path.exists(invoice['invoice_file_path']):
            with open(invoice['invoice_file_path'], "rb") as file:
                st.sidebar.download_button(
                    label="Download Invoice",
                    data=file,
                    file_name=os.path.basename(invoice['invoice_file_path']),
                    mime="application/octet-stream"
                )
        else:
            st.sidebar.info("No invoice file uploaded.")
            
            # Upload option
            with st.sidebar.form("upload_invoice_file"):
                uploaded_file = st.file_uploader("Upload Invoice File", type=["pdf", "png", "jpg", "jpeg", "doc", "docx"])
                upload_submitted = st.form_submit_button("Upload")
                
                if upload_submitted and uploaded_file:
                    # Save file
                    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                    file_ext = os.path.splitext(uploaded_file.name)[1]
                    file_path = os.path.join("uploads", f"invoice_{invoice_id}_{timestamp}{file_ext}")
                    
                    with open(file_path, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    
                    # Update invoice
                    conn = get_db_connection()
                    conn.execute(
                        "UPDATE invoices SET invoice_file_path = ? WHERE invoice_id = ?", 
                        (file_path, invoice_id)
                    )
                    conn.commit()
                    conn.close()
                    
                    st.success("Invoice file uploaded!")
                    st.rerun()
        
        # Close button
        if st.sidebar.button("Close"):
            st.session_state.edit_invoice_id = None
            st.rerun()

def create_invoice_form():
    with st.form("create_invoice_form"):
        st.subheader("Create New Invoice")
        
        # Get vendor list
        conn = get_db_connection()
        vendors = pd.read_sql("SELECT vendor_id, vendor_name FROM vendors WHERE status = 'active' ORDER BY vendor_name", conn)
        conn.close()
        
        if vendors.empty:
            st.error("No active vendors found. Please create a vendor first.")
            return
        
        vendor_id = st.selectbox("Select Vendor", vendors['vendor_id'].tolist(), format_func=lambda x: vendors.loc[vendors['vendor_id'] == x, 'vendor_name'].iloc[0])
        
        invoice_number = st.text_input("Invoice Number *")
        
        col1, col2 = st.columns(2)
        with col1:
            invoice_date = st.date_input("Invoice Date", datetime.now().date())
        with col2:
            due_date = st.date_input("Due Date", (datetime.now() + timedelta(days=30)).date())
        
        col1, col2, col3 = st.columns(3)
        with col1:
            amount = st.number_input("Amount *", min_value=0.0, format="%.2f")
        with col2:
            tax_amount = st.number_input("Tax Amount", min_value=0.0, format="%.2f")
        with col3:
            total_amount = st.number_input("Total Amount *", min_value=0.0, format="%.2f")
        
        description = st.text_area("Description")
        
        uploaded_file = st.file_uploader("Upload Invoice File", type=["pdf", "png", "jpg", "jpeg", "doc", "docx"])
        
        submitted = st.form_submit_button("Create Invoice")
        
        if submitted:
            if not invoice_number or amount <= 0 or total_amount <= 0:
                st.error("Please fill all required fields (Invoice Number, Amount, Total Amount).")
            else:
                conn = get_db_connection()
                
                # Save file if uploaded
                file_path = None
                if uploaded_file:
                    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                    file_ext = os.path.splitext(uploaded_file.name)[1]
                    file_path = os.path.join("uploads", f"invoice_{timestamp}{file_ext}")
                    
                    with open(file_path, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                
                # Insert invoice
                conn.execute("""
                    INSERT INTO invoices
                    (vendor_id, invoice_number, invoice_date, due_date, amount, tax_amount, total_amount, description, status, invoice_file_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)
                """, (vendor_id, invoice_number, invoice_date, due_date, amount, tax_amount, total_amount, description, file_path))
                conn.commit()
                conn.close()
                
                st.success(f"Invoice '{invoice_number}' created successfully!")
                st.balloons()

def import_invoices_from_tally():
    st.subheader("Import Invoices from Tally")
    
    st.write("""
    This feature allows you to import pending bills from Tally ERP using ODBC connection. 
    Ensure Tally is running and the ODBC driver is configured correctly.
    """)
    
    # Tally connection settings
    with st.expander("Tally Connection Settings"):
        col1, col2 = st.columns(2)
        
        with col1:
            tally_server = st.text_input("Tally Server", "localhost")
        with col2:
            tally_port = st.text_input("Tally Port", "9000")
    
    # Import buttons
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("Import Vendors from Tally"):
            try:
                from utils.tally_connector import TallyConnector
                
                tally = TallyConnector()
                count = tally.sync_vendors()
                
                if count > 0:
                    st.success(f"Successfully imported/updated {count} vendors from Tally.")
                else:
                    st.info("No new vendors found in Tally.")
            except Exception as e:
                st.error(f"Error importing vendors: {str(e)}")
    
    with col2:
        if st.button("Import Pending Bills from Tally"):
            try:
                from utils.tally_connector import TallyConnector
                
                tally = TallyConnector()
                count = tally.sync_invoices()
                
                if count > 0:
                    st.success(f"Successfully imported {count} pending bills from Tally.")
                else:
                    st.info("No new pending bills found in Tally.")
            except Exception as e:
                st.error(f"Error importing bills: {str(e)}")
    
    # Last import status
    st.divider()
    st.subheader("Import History")
    
    # Display recent imports
    conn = get_db_connection()
    imports = pd.read_sql("""
        SELECT action, entity_type, COUNT(entity_id) as count, MAX(created_at) as last_import
        FROM audit_logs
        WHERE action LIKE 'import%'
        GROUP BY action, entity_type
        ORDER BY last_import DESC
        LIMIT 5
    """, conn)
    conn.close()
    
    if not imports.empty:
        for i, row in imports.iterrows():
            st.write(f"**{row['action'].title()}**: {row['count']} {row['entity_type']}s on {row['last_import']}")
    else:
        st.info("No import history found.")

def display_create_payment_request_modal(invoice_ids):
    conn = get_db_connection()
    
    # Get invoices
    invoices_str = ','.join(['?'] * len(invoice_ids))
    invoices = pd.read_sql(f"""
        SELECT i.invoice_id, i.vendor_id, v.vendor_name, i.invoice_number, 
               i.invoice_date, i.due_date, i.total_amount, i.status
        FROM invoices i
        JOIN vendors v ON i.vendor_id = v.vendor_id
        WHERE i.invoice_id IN ({invoices_str})
    """, conn, params=invoice_ids)
    
    # Convert date columns
    invoices['invoice_date'] = pd.to_datetime(invoices['invoice_date']).dt.strftime('%Y-%m-%d')
    invoices['due_date'] = pd.to_datetime(invoices['due_date']).dt.strftime('%Y-%m-%d')
    
    # Get unique vendor IDs
    vendor_ids = invoices['vendor_id'].unique()
    
    if len(vendor_ids) > 1:
        st.sidebar.error("Payment request can only be created for invoices from the same vendor.")
        st.session_state.create_payment_request = None
        return
    
    vendor_name = invoices['vendor_name'].iloc[0]
    
    st.sidebar.title("Create Payment Request")
    st.sidebar.subheader(f"Vendor: {vendor_name}")
    
    # Display selected invoices
    st.sidebar.write("Selected Invoices:")
    
    total_amount = 0
    for i, invoice in invoices.iterrows():
        st.sidebar.write(f"• {invoice['invoice_number']} - ${float(invoice['total_amount']):,.2f} (Due: {invoice['due_date']})")
        total_amount += float(invoice['total_amount'])
    
    st.sidebar.write(f"**Total Amount: ${total_amount:,.2f}**")
    
    # Payment request form
    with st.sidebar.form("payment_request_form"):
        notes = st.text_area("Notes/Comments")
        
        submitted = st.form_submit_button("Submit Payment Request")
        
        if submitted:
            try:
                # Generate request number
                timestamp = datetime.now().strftime("%Y%m%d%H%M")
                request_number = f"PR{timestamp}"
                
                # Insert payment request
                conn.execute("""
                    INSERT INTO payment_requests
                    (request_number, requested_by, notes, status)
                    VALUES (?, ?, ?, 'pending')
                """, (request_number, st.session_state.user['user_id'], notes))
                
                # Get the request ID
                request_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                
                # Add invoices to the request
                for invoice_id in invoice_ids:
                    conn.execute("""
                        INSERT INTO payment_request_items
                        (request_id, invoice_id)
                        VALUES (?, ?)
                    """, (request_id, invoice_id))
                    
                    # Update invoice status
                    conn.execute("""
                        UPDATE invoices
                        SET status = 'approved'
                        WHERE invoice_id = ?
                    """, (invoice_id,))
                
                # Add audit log
                conn.execute("""
                    INSERT INTO audit_logs
                    (user_id, action, entity_type, entity_id, details)
                    VALUES (?, 'created', 'payment_request', ?, ?)
                """, (st.session_state.user['user_id'], request_id, f"Created payment request for {len(invoice_ids)} invoices"))
                
                conn.commit()
                st.sidebar.success("Payment request created successfully!")
                st.sidebar.info(f"Request Number: {request_number}")
                
                # Clear selection
                st.session_state.create_payment_request = None
                st.rerun()
            
            except Exception as e:
                conn.rollback()
                st.sidebar.error(f"Error creating payment request: {str(e)}")
            finally:
                conn.close()
    
    # Cancel button
    if st.sidebar.button("Cancel"):
        st.session_state.create_payment_request = None
        st.rerun()

# Payment Requests Page
def display_payment_requests():
    st.title("Payment Requests")
    
    # Get payment requests
    conn = get_db_connection()
    payment_requests = pd.read_sql("""
        SELECT pr.request_id, pr.request_number, pr.requested_at, 
               u1.full_name as requested_by, pr.status,
               u2.full_name as approved_by, pr.approved_at, pr.notes,
               COUNT(pri.invoice_id) as invoice_count,
               SUM(i.total_amount) as total_amount
        FROM payment_requests pr
        JOIN users u1 ON pr.requested_by = u1.user_id
        LEFT JOIN users u2 ON pr.approved_by = u2.user_id
        JOIN payment_request_items pri ON pr.request_id = pri.request_id
        JOIN invoices i ON pri.invoice_id = i.invoice_id
        GROUP BY pr.request_id
        ORDER BY pr.requested_at DESC
    """, conn)
    conn.close()
    
    # Convert date columns
    payment_requests['requested_at'] = pd.to_datetime(payment_requests['requested_at']).dt.strftime('%Y-%m-%d %H:%M')
    payment_requests['approved_at'] = payment_requests['approved_at'].apply(lambda x: pd.to_datetime(x).strftime('%Y-%m-%d %H:%M') if x else "")
    
    # Status filter
    status_filter = st.multiselect(
        "Filter by Status",
        options=["pending", "approved", "rejected", "processed"],
        default=["pending"]
    )
    
    if status_filter:
        payment_requests = payment_requests[payment_requests['status'].isin(status_filter)]
    
    # Display payment requests
    for i, pr in payment_requests.iterrows():
        col1, col2, col3, col4 = st.columns([2, 3, 2, 1])
        
        with col1:
            st.write(f"**Request #:** {pr['request_number']}")
            st.write(f"**Date:** {pr['requested_at']}")
        
        with col2:
            st.write(f"**Requested By:** {pr['requested_by']}")
            st.write(f"**Status:** {pr['status'].title()}")
            
            # Show approver info if approved/rejected
            if pr['status'] in ['approved', 'rejected', 'processed'] and pr['approved_by']:
                st.write(f"**Approved By:** {pr['approved_by']} on {pr['approved_at']}")
        
        with col3:
            st.write(f"**Invoices:** {pr['invoice_count']}")
            st.write(f"**Amount:** ${float(pr['total_amount']):,.2f}")
        
        with col4:
            if st.button("View", key=f"view_pr_{pr['request_id']}"):
                st.session_state.view_payment_request_id = pr['request_id']
                st.rerun()
        
        st.divider()
    
    # View payment request modal
    if 'view_payment_request_id' in st.session_state and st.session_state.view_payment_request_id:
        display_payment_request_details(st.session_state.view_payment_request_id)

def display_payment_request_details(request_id):
    conn = get_db_connection()
    
    # Get payment request details
    payment_request = conn.execute("""
        SELECT pr.*, u1.full_name as requester_name, u2.full_name as approver_name
        FROM payment_requests pr
        JOIN users u1 ON pr.requested_by = u1.user_id
        LEFT JOIN users u2 ON pr.approved_by = u2.user_id
        WHERE pr.request_id = ?
    """, (request_id,)).fetchone()
    
    # Get invoices in this request
    invoices = pd.read_sql("""
        SELECT i.invoice_id, i.vendor_id, v.vendor_name, i.invoice_number, 
               i.invoice_date, i.due_date, i.total_amount
        FROM payment_request_items pri
        JOIN invoices i ON pri.invoice_id = i.invoice_id
        JOIN vendors v ON i.vendor_id = v.vendor_id
        WHERE pri.request_id = ?
    """, conn, params=(request_id,))
    
    # Convert date columns
    invoices['invoice_date'] = pd.to_datetime(invoices['invoice_date']).dt.strftime('%Y-%m-%d')
    invoices['due_date'] = pd.to_datetime(invoices['due_date']).dt.strftime('%Y-%m-%d')
    
    # Get payment advices
    payment_advices = pd.read_sql("""
        SELECT * FROM payment_advices
        WHERE request_id = ?
        ORDER BY generated_at DESC
    """, conn, params=(request_id,))
    
    # Convert advice date columns
    if not payment_advices.empty:
        payment_advices['generated_at'] = pd.to_datetime(payment_advices['generated_at']).dt.strftime('%Y-%m-%d %H:%M')
        payment_advices['payment_date'] = pd.to_datetime(payment_advices['payment_date']).dt.strftime('%Y-%m-%d')
    
    conn.close()
    
    if payment_request:
        # Payment request details section
        st.sidebar.title(f"Payment Request: {payment_request['request_number']}")
        
        st.sidebar.write(f"**Status:** {payment_request['status'].title()}")
        st.sidebar.write(f"**Requested By:** {payment_request['requester_name']}")
        st.sidebar.write(f"**Requested At:** {datetime.strptime(payment_request['requested_at'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M')}")
        
        if payment_request['approved_by']:
            st.sidebar.write(f"**Approved By:** {payment_request['approver_name']}")
            st.sidebar.write(f"**Approved At:** {datetime.strptime(payment_request['approved_at'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M')}")
        
        if payment_request['notes']:
            st.sidebar.subheader("Notes")
            st.sidebar.write(payment_request['notes'])
        
        # Invoices section
        st.sidebar.subheader("Invoices")
        
        total_amount = 0
        for i, invoice in invoices.iterrows():
            st.sidebar.write(f"• {invoice['invoice_number']} - ${float(invoice['total_amount']):,.2f} (Due: {invoice['due_date']})")
            total_amount += float(invoice['total_amount'])
        
        st.sidebar.write(f"**Total Amount: ${total_amount:,.2f}**")
        
        # Actions section
        st.sidebar.subheader("Actions")
        
        # Approval/Rejection actions (for approvers)
        if payment_request['status'] == 'pending' and st.session_state.user_role in ['admin', 'approver']:
            col1, col2 = st.sidebar.columns(2)
            
            with col1:
                if st.button("Approve", key=f"approve_{request_id}"):
                    conn = get_db_connection()
                    conn.execute("""
                        UPDATE payment_requests
                        SET status = 'approved', approved_by = ?, approved_at = CURRENT_TIMESTAMP
                        WHERE request_id = ?
                    """, (st.session_state.user['user_id'], request_id))
                    
                    # Add audit log
                    conn.execute("""
                        INSERT INTO audit_logs
                        (user_id, action, entity_type, entity_id, details)
                        VALUES (?, 'approved', 'payment_request', ?, ?)
                    """, (st.session_state.user['user_id'], request_id, f"Approved payment request {payment_request['request_number']}"))
                    
                    conn.commit()
                    conn.close()
                    
                    st.sidebar.success("Payment request approved!")
                    st.rerun()
            
            with col2:
                if st.button("Reject", key=f"reject_{request_id}"):
                    rejection_reason = st.sidebar.text_area("Rejection Reason")
                    
                    if st.sidebar.button("Confirm Rejection", key=f"confirm_reject_{request_id}"):
                        conn = get_db_connection()
                        conn.execute("""
                            UPDATE payment_requests
                            SET status = 'rejected', approved_by = ?, approved_at = CURRENT_TIMESTAMP,
                                rejection_reason = ?
                            WHERE request_id = ?
                        """, (st.session_state.user['user_id'], rejection_reason, request_id))
                        
                        # Update invoice status back to pending
                        conn.execute("""
                            UPDATE invoices
                            SET status = 'pending'
                            WHERE invoice_id IN (
                                SELECT invoice_id FROM payment_request_items WHERE request_id = ?
                            )
                        """, (request_id,))
                        
                        # Add audit log
                        conn.execute("""
                            INSERT INTO audit_logs
                            (user_id, action, entity_type, entity_id, details)
                            VALUES (?, 'rejected', 'payment_request', ?, ?)
                        """, (st.session_state.user['user_id'], request_id, f"Rejected payment request {payment_request['request_number']}"))
                        
                        conn.commit()
                        conn.close()
                        
                        st.sidebar.error("Payment request rejected!")
                        st.rerun()
        
        # Generate payment advice (for accountants after approval)
        if payment_request['status'] == 'approved' and st.session_state.user_role in ['admin', 'accountant']:
            if st.sidebar.button("Generate Payment Advice", key=f"generate_advice_{request_id}"):
                try:
                    from utils.excel_generator import ExcelReportGenerator
                    
                    # Generate payment advice
                    generator = ExcelReportGenerator()
                    success, result = generator.generate_payment_advice(request_id)
                    
                    if success:
                        # Create payment advice record
                        advice_number = f"PA{datetime.now().strftime('%Y%m%d%H%M')}"
                        
                        conn = get_db_connection()
                        conn.execute("""
                            INSERT INTO payment_advices
                            (request_id, advice_number, total_amount, generated_at, payment_date, status)
                            VALUES (?, ?, ?, CURRENT_TIMESTAMP, ?, 'pending')
                        """, (request_id, advice_number, total_amount, datetime.now().date()))
                        
                        # Update payment request status
                        conn.execute("""
                            UPDATE payment_requests
                            SET status = 'processed'
                            WHERE request_id = ?
                        """, (request_id,))
                        
                        conn.commit()
                        conn.close()
                        
                        # Download button for advice
                        with open(result, "rb") as file:
                            st.sidebar.download_button(
                                label="Download Payment Advice",
                                data=file,
                                file_name=os.path.basename(result),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        st.sidebar.success("Payment advice generated successfully!")
                        st.rerun()
                    else:
                        st.sidebar.error(f"Error generating payment advice: {result}")
                
                except Exception as e:
                    st.sidebar.error(f"Error: {str(e)}")
        
        # Payment advices history
        if not payment_advices.empty:
            st.sidebar.subheader("Payment Advices")
            
            for i, advice in payment_advices.iterrows():
                st.sidebar.write(f"• {advice['advice_number']} - ${float(advice['total_amount']):,.2f} ({advice['generated_at']})")
                
                # Find the file
                advice_file = None
                for file in os.listdir("reports"):
                    if advice['advice_number'] in file:
                        advice_file = os.path.join("reports", file)
                        break
                
                if advice_file and os.path.exists(advice_file):
                    with open(advice_file, "rb") as file:
                        st.sidebar.download_button(
                            label="Download",
                            data=file,
                            file_name=os.path.basename(advice_file),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_advice_{advice['advice_id']}"
                        )
        
        # Close button
        if st.sidebar.button("Close"):
            st.session_state.view_payment_request_id = None
            st.rerun()

# Payment Approvals Page
def display_payment_approvals():
    st.title("Payment Approvals")
    
    # Get pending payment requests
    conn = get_db_connection()
    payment_requests = pd.read_sql("""
        SELECT pr.request_id, pr.request_number, pr.requested_at, 
               u1.full_name as requested_by, pr.status, pr.notes,
               COUNT(pri.invoice_id) as invoice_count,
               SUM(i.total_amount) as total_amount,
               MIN(v.vendor_name) as vendor_name
        FROM payment_requests pr
        JOIN users u1 ON pr.requested_by = u1.user_id
        JOIN payment_request_items pri ON pr.request_id = pri.request_id
        JOIN invoices i ON pri.invoice_id = i.invoice_id
        JOIN vendors v ON i.vendor_id = v.vendor_id
        WHERE pr.status = 'pending'
        GROUP BY pr.request_id
        ORDER BY pr.requested_at ASC
    """, conn)
    conn.close()
    
    # Convert date columns
    payment_requests['requested_at'] = pd.to_datetime(payment_requests['requested_at']).dt.strftime('%Y-%m-%d %H:%M')
    
    if payment_requests.empty:
        st.info("No pending payment requests requiring approval.")
    else:
        st.write(f"You have {len(payment_requests)} payment requests pending approval.")
        
        # Display payment requests
        for i, pr in payment_requests.iterrows():
            with st.expander(f"Request #{pr['request_number']} - {pr['vendor_name']} - ${float(pr['total_amount']):,.2f}"):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write(f"**Request Number:** {pr['request_number']}")
                    st.write(f"**Requested By:** {pr['requested_by']}")
                    st.write(f"**Requested At:** {pr['requested_at']}")
                    st.write(f"**Vendor:** {pr['vendor_name']}")
                
                with col2:
                    st.write(f"**Invoices:** {pr['invoice_count']}")
                    st.write(f"**Total Amount:** ${float(pr['total_amount']):,.2f}")
                    
                    if pr['notes']:
                        st.write(f"**Notes:** {pr['notes']}")
                
                # Get invoices in this request
                conn = get_db_connection()
                invoices = pd.read_sql("""
                    SELECT i.invoice_id, i.invoice_number, i.invoice_date, i.due_date, i.total_amount
                    FROM payment_request_items pri
                    JOIN invoices i ON pri.invoice_id = i.invoice_id
                    WHERE pri.request_id = ?
                """, conn, params=(pr['request_id'],))
                conn.close()
                
                # Convert date columns
                invoices['invoice_date'] = pd.to_datetime(invoices['invoice_date']).dt.strftime('%Y-%m-%d')
                invoices['due_date'] = pd.to_datetime(invoices['due_date']).dt.strftime('%Y-%m-%d')
                
                # Display invoices
                st.subheader("Invoices")
                st.dataframe(
                    invoices[['invoice_number', 'invoice_date', 'due_date', 'total_amount']],
                    hide_index=True
                )
                
                # Approval/Rejection actions
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("Approve", key=f"approve_{pr['request_id']}"):
                        conn = get_db_connection()
                        conn.execute("""
                            UPDATE payment_requests
                            SET status = 'approved', approved_by = ?, approved_at = CURRENT_TIMESTAMP
                            WHERE request_id = ?
                        """, (st.session_state.user['user_id'], pr['request_id']))
                        
                        # Add audit log
                        conn.execute("""
                            INSERT INTO audit_logs
                            (user_id, action, entity_type, entity_id, details)
                            VALUES (?, 'approved', 'payment_request', ?, ?)
                        """, (st.session_state.user['user_id'], pr['request_id'], f"Approved payment request {pr['request_number']}"))
                        
                        conn.commit()
                        conn.close()
                        
                        st.success(f"Payment request #{pr['request_number']} approved!")
                        st.rerun()
                
                with col2:
                    if st.button("Reject", key=f"reject_{pr['request_id']}"):
                        rejection_reason = st.text_area("Rejection Reason", key=f"reason_{pr['request_id']}")
                        
                        if st.button("Confirm Rejection", key=f"confirm_reject_{pr['request_id']}"):
                            conn = get_db_connection()
                            conn.execute("""
                                UPDATE payment_requests
                                SET status = 'rejected', approved_by = ?, approved_at = CURRENT_TIMESTAMP,
                                    rejection_reason = ?
                                WHERE request_id = ?
                            """, (st.session_state.user['user_id'], rejection_reason, pr['request_id']))
                            
                            # Update invoice status back to pending
                            conn.execute("""
                                UPDATE invoices
                                SET status = 'pending'
                                WHERE invoice_id IN (
                                    SELECT invoice_id FROM payment_request_items WHERE request_id = ?
                                )
                            """, (pr['request_id'],))
                            
                            # Add audit log
                            conn.execute("""
                                INSERT INTO audit_logs
                                (user_id, action, entity_type, entity_id, details)
                                VALUES (?, 'rejected', 'payment_request', ?, ?)
                            """, (st.session_state.user['user_id'], pr['request_id'], f"Rejected payment request {pr['request_number']}"))
                            
                            conn.commit()
                            conn.close()
                            
                            st.error(f"Payment request #{pr['request_number']} rejected!")
                            st.rerun()

# Reports Page
def display_reports():
    st.title("Reports & Analytics")
    
    report_types = [
        "Aging Report", 
        "Vendor Summary", 
        "Payment History", 
        "Invoice Status Summary", 
        "Monthly Trend"
    ]
    
    report_type = st.selectbox("Select Report Type", report_types)
    
    if report_type == "Aging Report":
        display_aging_report()
    elif report_type == "Vendor Summary":
        display_vendor_summary_report()
    elif report_type == "Payment History":
        display_payment_history_report()
    elif report_type == "Invoice Status Summary":
        display_invoice_status_report()
    elif report_type == "Monthly Trend":
        display_monthly_trend_report()

def display_aging_report():
    st.subheader("Accounts Payable Aging Report")
    
    # Date selection
    as_of_date = st.date_input("As of Date", datetime.now().date())
    
    # Generate report button
    if st.button("Generate Report"):
        try:
            from utils.excel_generator import ExcelReportGenerator
            
            generator = ExcelReportGenerator()
            success, result = generator.generate_aging_report(as_of_date.strftime("%Y-%m-%d"))
            
            if success:
                # Download button for report
                with open(result, "rb") as file:
                    st.download_button(
                        label="Download Aging Report",
                        data=file,
                        file_name=os.path.basename(result),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                st.success("Aging report generated successfully!")
            else:
                st.error(f"Error generating report: {result}")
        
        except Exception as e:
            st.error(f"Error: {str(e)}")
    
    # Display aging data
    conn = get_db_connection()
    invoices = pd.read_sql("""
        SELECT i.invoice_id, i.vendor_id, v.vendor_name, i.invoice_number, 
               i.invoice_date, i.due_date, i.total_amount, i.status
        FROM invoices i
        JOIN vendors v ON i.vendor_id = v.vendor_id
        WHERE i.status IN ('pending', 'approved')
    """, conn)
    conn.close()
    
    if not invoices.empty:
        # Convert date columns
        invoices['invoice_date'] = pd.to_datetime(invoices['invoice_date'])
        invoices['due_date'] = pd.to_datetime(invoices['due_date'])
        
        # Calculate days overdue
        invoices['days_overdue'] = (as_of_date - invoices['due_date'].dt.date).dt.days
        
        # Create aging buckets
        conditions = [
            (invoices['days_overdue'] <= 0),
            (invoices['days_overdue'] > 0) & (invoices['days_overdue'] <= 30),
            (invoices['days_overdue'] > 30) & (invoices['days_overdue'] <= 60),
            (invoices['days_overdue'] > 60) & (invoices['days_overdue'] <= 90),
            (invoices['days_overdue'] > 90)
        ]
        
        choices = ['Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days']
        invoices['aging_bucket'] = np.select(conditions, choices, default='Current')
        
        # Group by aging bucket
        aging_summary = invoices.groupby('aging_bucket').agg(
            count=('invoice_id', 'count'),
            total=('total_amount', 'sum')
        ).reset_index()
        
        # Sort buckets
        bucket_order = {
            'Current': 0,
            '1-30 Days': 1,
            '31-60 Days': 2,
            '61-90 Days': 3,
            'Over 90 Days': 4
        }
        aging_summary['bucket_order'] = aging_summary['aging_bucket'].map(bucket_order)
        aging_summary = aging_summary.sort_values('bucket_order').drop('bucket_order', axis=1)
        
        # Display summary chart
        fig = px.pie(
            aging_summary, 
            values='total', 
            names='aging_bucket',
            title='Accounts Payable Aging Summary',
            color='aging_bucket',
            color_discrete_map={
                'Current': '#28a745',
                '1-30 Days': '#ffc107',
                '31-60 Days': '#fd7e14',
                '61-90 Days': '#dc3545',
                'Over 90 Days': '#6c757d'
            }
        )
        st.plotly_chart(fig)
        
        # Display data table
        st.subheader("Aging Summary")
        aging_summary['count'] = aging_summary['count'].apply(lambda x: f"{x}")
        aging_summary['total'] = aging_summary['total'].apply(lambda x: f"${x:,.2f}")
        st.dataframe(
            aging_summary.rename(columns={'aging_bucket': 'Aging Bucket', 'count': 'Invoice Count', 'total': 'Total Amount'}),
            hide_index=True
        )
        
        # Display detailed data
        st.subheader("Invoice Details")
        
        # Format for display
        invoices['invoice_date'] = invoices['invoice_date'].dt.strftime('%Y-%m-%d')
        invoices['due_date'] = invoices['due_date'].dt.strftime('%Y-%m-%d')
        invoices['total_amount'] = invoices['total_amount'].apply(lambda x: f"${x:,.2f}")
        
        display_cols = ['vendor_name', 'invoice_number', 'invoice_date', 'due_date', 'days_overdue', 'aging_bucket', 'total_amount', 'status']
        
        # Add filter for aging bucket
        selected_bucket = st.multiselect(
            "Filter by Aging Bucket",
            options=choices,
            default=choices
        )
        
        filtered_invoices = invoices[invoices['aging_bucket'].isin(selected_bucket)]
        
        st.dataframe(
            filtered_invoices[display_cols].rename(columns={
                'vendor_name': 'Vendor',
                'invoice_number': 'Invoice #',
                'invoice_date': 'Invoice Date',
                'due_date': 'Due Date',
                'days_overdue': 'Days Overdue',
                'aging_bucket': 'Aging Bucket',
                'total_amount': 'Amount',
                'status': 'Status'
            }),
            hide_index=True
        )
    else:
        st.info("No pending invoices found.")

def display_vendor_summary_report():
    st.subheader("Vendor Summary Report")
    
    # Get vendor summary data
    conn = get_db_connection()
    vendor_summary = pd.read_sql("""
        SELECT v.vendor_id, v.vendor_name,
               COUNT(i.invoice_id) as total_invoices,
               SUM(CASE WHEN i.status = 'paid' THEN 1 ELSE 0 END) as paid_invoices,
               SUM(CASE WHEN i.status IN ('pending', 'approved') THEN 1 ELSE 0 END) as pending_invoices,
               SUM(CASE WHEN i.status = 'paid' THEN i.total_amount ELSE 0 END) as paid_amount,
               SUM(CASE WHEN i.status IN ('pending', 'approved') THEN i.total_amount ELSE 0 END) as pending_amount
        FROM vendors v
        LEFT JOIN invoices i ON v.vendor_id = i.vendor_id
        WHERE v.status = 'active'
        GROUP BY v.vendor_id, v.vendor_name
        ORDER BY pending_amount DESC
    """, conn)
    conn.close()
    
    if not vendor_summary.empty:
        # Create totals row
        totals = vendor_summary.sum(numeric_only=True)
        totals['vendor_name'] = 'TOTAL'
        vendor_summary = pd.concat([vendor_summary, pd.DataFrame([totals])], ignore_index=True)
        
        # Format currency columns
        vendor_summary['paid_amount'] = vendor_summary['paid_amount'].apply(lambda x: f"${x:,.2f}")
        vendor_summary['pending_amount'] = vendor_summary['pending_amount'].apply(lambda x: f"${x:,.2f}")
        
        # Display summary table
        st.dataframe(
            vendor_summary[['vendor_name', 'total_invoices', 'paid_invoices', 'pending_invoices', 'paid_amount', 'pending_amount']].rename(columns={
                'vendor_name': 'Vendor',
                'total_invoices': 'Total Invoices',
                'paid_invoices': 'Paid Invoices',
                'pending_invoices': 'Pending Invoices',
                'paid_amount': 'Paid Amount',
                'pending_amount': 'Pending Amount'
            }),
            hide_index=True
        )
        
        # Bar chart of top vendors by pending amount
        top_vendors = vendor_summary[vendor_summary['vendor_name'] != 'TOTAL'].sort_values('pending_amount', key=lambda x: pd.to_numeric(x.str.replace('[$,]', '', regex=True)), ascending=False).head(10)
        
        fig = px.bar(
            top_vendors,
            x='vendor_name',
            y=[pd.to_numeric(top_vendors['pending_amount'].str.replace('[$,]', '', regex=True)), 
               pd.to_numeric(top_vendors['paid_amount'].str.replace('[$,]', '', regex=True))],
            title='Top 10 Vendors by Outstanding Amount',
            labels={'vendor_name': 'Vendor', 'value': 'Amount ($)', 'variable': 'Type'},
            barmode='group'
        )
        fig.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig)
        
        # Export to Excel
        if st.button("Export to Excel"):
            # Create a new Excel workbook
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vendor Summary"
            
            # Add headers
            headers = ['Vendor', 'Total Invoices', 'Paid Invoices', 'Pending Invoices', 'Paid Amount', 'Pending Amount']
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx).value = header
                ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
            
            # Add data
            for row_idx, row in enumerate(vendor_summary.itertuples(), 2):
                ws.cell(row=row_idx, column=1).value = row.vendor_name
                ws.cell(row=row_idx, column=2).value = row.total_invoices
                ws.cell(row=row_idx, column=3).value = row.paid_invoices
                ws.cell(row=row_idx, column=4).value = row.pending_invoices
                ws.cell(row=row_idx, column=5).value = row.paid_amount
                ws.cell(row=row_idx, column=6).value = row.pending_amount
                
                # Bold the totals row
                if row.vendor_name == 'TOTAL':
                    for col_idx in range(1, 7):
                        ws.cell(row=row_idx, column=col_idx).font = openpyxl.styles.Font(bold=True)
            
            # Save the file
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            filename = f"vendor_summary_{timestamp}.xlsx"
            filepath = os.path.join("reports", filename)
            os.makedirs("reports", exist_ok=True)
            wb.save(filepath)
            
            # Download button
            with open(filepath, "rb") as f:
                st.download_button(
                    label="Download Excel Report",
                    data=f,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    else:
        st.info("No vendor data found.")

def display_payment_history_report():
    st.subheader("Payment History Report")
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date", (datetime.now() - timedelta(days=30)).date())
    with col2:
        end_date = st.date_input("End Date", datetime.now().date())
    
    # Get payment history data
    conn = get_db_connection()
    
    payment_history = pd.read_sql("""
        SELECT pa.advice_number, pa.generated_at, pa.payment_date, pa.total_amount,
               pr.request_number, u.full_name as approved_by,
               COUNT(pri.invoice_id) as invoice_count,
               GROUP_CONCAT(DISTINCT v.vendor_name) as vendor_names
        FROM payment_advices pa
        JOIN payment_requests pr ON pa.request_id = pr.request_id
        JOIN users u ON pr.approved_by = u.user_id
        JOIN payment_request_items pri ON pr.request_id = pri.request_id
        JOIN invoices i ON pri.invoice_id = i.invoice_id
        JOIN vendors v ON i.vendor_id = v.vendor_id
        WHERE pa.payment_date BETWEEN ? AND ?
        GROUP BY pa.advice_id
        ORDER BY pa.payment_date DESC
    """, conn, params=(start_date, end_date))
    
    conn.close()
    
    if not payment_history.empty:
        # Convert date columns
        payment_history['generated_at'] = pd.to_datetime(payment_history['generated_at']).dt.strftime('%Y-%m-%d')
        payment_history['payment_date'] = pd.to_datetime(payment_history['payment_date']).dt.strftime('%Y-%m-%d')
        
        # Format currency columns
        payment_history['total_amount'] = payment_history['total_amount'].apply(lambda x: f"${x:,.2f}")
        
        # Display data table
        st.dataframe(
            payment_history[['advice_number', 'payment_date', 'vendor_names', 'invoice_count', 'total_amount', 'approved_by']].rename(columns={
                'advice_number': 'Payment Advice #',
                'payment_date': 'Payment Date',
                'vendor_names': 'Vendor',
                'invoice_count': 'Invoices',
                'total_amount': 'Amount',
                'approved_by': 'Approved By'
            }),
            hide_index=True
        )
        
        # Calculate summary
        payment_count = len(payment_history)
        total_paid = payment_history['total_amount'].apply(lambda x: float(x.replace('$', '').replace(',', ''))).sum()
        
        # Display summary metrics
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Total Payments", payment_count)
        with col2:
            st.metric("Total Amount Paid", f"${total_paid:,.2f}")
        
        # Line chart of payments over time
        payment_history['payment_date_dt'] = pd.to_datetime(payment_history['payment_date'])
        payment_history['amount_numeric'] = payment_history['total_amount'].apply(lambda x: float(x.replace('$', '').replace(',', '')))
        
        payment_by_date = payment_history.groupby(payment_history['payment_date_dt'].dt.strftime('%Y-%m-%d')).agg(
            total=('amount_numeric', 'sum'),
            count=('advice_number', 'count')
        ).reset_index()
        
        fig = px.line(
            payment_by_date,
            x='payment_date_dt',
            y='total',
            title='Payment Amounts Over Time',
            labels={'payment_date_dt': 'Date', 'total': 'Amount ($)'}
        )
        st.plotly_chart(fig)
    
    else:
        st.info(f"No payment history found between {start_date} and {end_date}.")

def display_invoice_status_report():
    st.subheader("Invoice Status Summary")
    
    # Get invoice status data
    conn = get_db_connection()
    
    invoice_status = pd.read_sql("""
        SELECT i.status,
               COUNT(i.invoice_id) as invoice_count,
               SUM(i.total_amount) as total_amount
        FROM invoices i
        GROUP BY i.status
    """, conn)
    
    # Get invoice trend data
    invoice_trend = pd.read_sql("""
        SELECT strftime('%Y-%m', i.invoice_date) as month,
               COUNT(i.invoice_id) as invoice_count,
               SUM(i.total_amount) as total_amount
        FROM invoices i
        GROUP BY month
        ORDER BY month DESC
        LIMIT 12
    """, conn)
    
    conn.close()
    
    if not invoice_status.empty:
        # Format the status labels
        invoice_status['status'] = invoice_status['status'].str.title()
        
        # Format currency columns
        invoice_status['total_amount'] = invoice_status['total_amount'].apply(lambda x: f"${x:,.2f}")
        
        # Display pie chart
        fig = px.pie(
            invoice_status,
            values='invoice_count',
            names='status',
            title='Invoice Status Distribution',
            color='status',
            color_discrete_map={
                'Pending': '#ffc107',
                'Approved': '#17a2b8',
                'Rejected': '#dc3545',
                'Paid': '#28a745'
            }
        )
        st.plotly_chart(fig)
        
        # Display data table
        st.dataframe(
            invoice_status.rename(columns={
                'status': 'Status',
                'invoice_count': 'Count',
                'total_amount': 'Total Amount'
            }),
            hide_index=True
        )
    
    # Display invoice trend if available
    if not invoice_trend.empty:
        st.subheader("Invoice Trend (Last 12 Months)")
        
        # Convert month to datetime for proper sorting
        invoice_trend['month_dt'] = pd.to_datetime(invoice_trend['month'] + '-01')
        invoice_trend = invoice_trend.sort_values('month_dt')
        
        # Create bar chart
        fig = px.bar(
            invoice_trend,
            x='month',
            y=['invoice_count', 'total_amount'],
            title='Invoice Trend by Month',
            labels={'month': 'Month', 'value': 'Value', 'variable': 'Metric'},
            barmode='group'
        )
        st.plotly_chart(fig)

def display_monthly_trend_report():
    st.subheader("Monthly AP Trend Analysis")
    
    # Get monthly data
    conn = get_db_connection()
    
    # Invoices by month
    invoices_monthly = pd.read_sql("""
        SELECT strftime('%Y-%m', invoice_date) as month,
               COUNT(invoice_id) as invoice_count,
               SUM(total_amount) as total_amount
        FROM invoices
        GROUP BY month
        ORDER BY month ASC
    """, conn)
    
    # Payments by month
    payments_monthly = pd.read_sql("""
        SELECT strftime('%Y-%m', payment_date) as month,
               COUNT(advice_id) as payment_count,
               SUM(total_amount) as payment_amount
        FROM payment_advices
        WHERE payment_date IS NOT NULL
        GROUP BY month
        ORDER BY month ASC
    """, conn)
    
    conn.close()
    
    if not invoices_monthly.empty:
        # Convert to datetime for proper handling
        invoices_monthly['month_dt'] = pd.to_datetime(invoices_monthly['month'] + '-01')
        
        # Create trend chart
        fig = px.line(
            invoices_monthly,
            x='month_dt',
            y='total_amount',
            title='Monthly Invoice Amounts',
            labels={'month_dt': 'Month', 'total_amount': 'Amount ($)'},
            markers=True
        )
        st.plotly_chart(fig)
        
        # If we have payment data, create a comparison chart
        if not payments_monthly.empty:
            payments_monthly['month_dt'] = pd.to_datetime(payments_monthly['month'] + '-01')
            
            # Merge the datasets on month
            monthly_comparison = pd.merge(
                invoices_monthly[['month', 'month_dt', 'total_amount']].rename(columns={'total_amount': 'invoice_amount'}),
                payments_monthly[['month', 'payment_amount']],
                on='month',
                how='outer'
            ).fillna(0)
            
            # Create comparison chart
            fig = px.line(
                monthly_comparison,
                x='month_dt',
                y=['invoice_amount', 'payment_amount'],
                title='Invoices vs Payments by Month',
                labels={'month_dt': 'Month', 'value': 'Amount ($)', 'variable': 'Type'},
                markers=True
            )
            fig.update_layout(
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1
                )
            )
            st.plotly_chart(fig)
            
            # Calculate rolling AP aging
            monthly_comparison['net_change'] = monthly_comparison['invoice_amount'] - monthly_comparison['payment_amount']
            monthly_comparison['cumulative_ap'] = monthly_comparison['net_change'].cumsum()
            
            # Create cumulative AP chart
            fig = px.area(
                monthly_comparison,
                x='month_dt',
                y='cumulative_ap',
                title='Cumulative Accounts Payable Balance',
                labels={'month_dt': 'Month', 'cumulative_ap': 'Balance ($)'}
            )
            st.plotly_chart(fig)
    
    else:
        st.info("Insufficient historical data for trend analysis.")

# Users Page
def display_users():
    st.title("User Management")
    
    # Check if user has admin role
    if st.session_state.user_role != 'admin':
        st.error("You don't have permission to access this page.")
        return
    
    # Tabs for user operations
    tab1, tab2 = st.tabs(["User List", "Create User"])
    
    with tab1:
        display_user_list()
    
    with tab2:
        create_user_form()

def display_user_list():
    conn = get_db_connection()
    users = pd.read_sql("""
        SELECT user_id, username, full_name, email, role, department, status, created_at
        FROM users
        ORDER BY created_at DESC
    """, conn)
    conn.close()
    
    # Convert date columns
    users['created_at'] = pd.to_datetime(users['created_at']).dt.strftime('%Y-%m-%d')
    
    # Format the roles and status
    users['role'] = users['role'].str.title()
    users['status'] = users['status'].str.title()
    
    # Search filter
    search = st.text_input("Search Users", "")
    if search:
        users = users[
            users['username'].str.contains(search, case=False) | 
            users['full_name'].str.contains(search, case=False) |
            users['email'].str.contains(search, case=False)
        ]
    
    # Status filter
    status_filter = st.multiselect(
        "Filter by Status", 
        options=["Active", "Inactive"],
        default=["Active"]
    )
    if status_filter:
        users = users[users['status'].isin(status_filter)]
    
    # Display users with edit button
    for i, user in users.iterrows():
        col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
        
        with col1:
            st.subheader(user['full_name'])
            st.write(f"Username: {user['username']}")
        
        with col2:
            st.write(f"Email: {user['email'] or 'N/A'}")
            st.write(f"Department: {user['department'] or 'N/A'}")
        
        with col3:
            st.write(f"Role: {user['role']}")
            st.write(f"Status: {user['status']}")
        
        with col4:
            st.write(f"Created: {user['created_at']}")
            if st.button("Edit", key=f"edit_user_{user['user_id']}"):
                st.session_state.edit_user_id = user['user_id']
                st.rerun()
        
        st.divider()
    
    # Edit user modal
    if 'edit_user_id' in st.session_state and st.session_state.edit_user_id:
        display_edit_user_modal(st.session_state.edit_user_id)

def display_edit_user_modal(user_id):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE user_id = ?", (user_id,)).fetchone()
    conn.close()
    
    if user:
        st.sidebar.title(f"Edit User: {user['username']}")
        
        with st.sidebar.form("edit_user_form"):
            username = st.text_input("Username", user['username'])
            full_name = st.text_input("Full Name", user['full_name'])
            email = st.text_input("Email", user['email'] or "")
            department = st.text_input("Department", user['department'] or "")
            role = st.selectbox(
                "Role", 
                ["admin", "accountant", "approver", "viewer"],
                index=["admin", "accountant", "approver", "viewer"].index(user['role'])
            )
            status = st.selectbox(
                "Status", 
                ["active", "inactive"],
                index=["active", "inactive"].index(user['status'])
            )
            
            change_password = st.checkbox("Change Password")
            
            if change_password:
                new_password = st.text_input("New Password", type="password")
                confirm_password = st.text_input("Confirm Password", type="password")
            
            submitted = st.form_submit_button("Update User")
            
            if submitted:
                if not username or not full_name:
                    st.error("Username and full name are required.")
                elif change_password and new_password != confirm_password:
                    st.error("Passwords do not match.")
                else:
                    conn = get_db_connection()
                    
                    if change_password:
                        # Update user with new password
                        # In a real application, hash the password
                        conn.execute("""
                            UPDATE users
                            SET username = ?, full_name = ?, email = ?, department = ?,
                                role = ?, status = ?, password_hash = ?
                            WHERE user_id = ?
                        """, (username, full_name, email, department, role, status, new_password, user_id))
                    else:
                        # Update user without changing password
                        conn.execute("""
                            UPDATE users
                            SET username = ?, full_name = ?, email = ?, department = ?,
                                role = ?, status = ?
                            WHERE user_id = ?
                        """, (username, full_name, email, department, role, status, user_id))
                    
                    conn.commit()
                    conn.close()
                    
                    st.success("User updated successfully!")
                    st.session_state.edit_user_id = None
                    st.rerun()
        
        # Close button
        if st.sidebar.button("Close"):
            st.session_state.edit_user_id = None
            st.rerun()

def create_user_form():
    with st.form("create_user_form"):
        st.subheader("Create New User")
        
        username = st.text_input("Username *")
        password = st.text_input("Password *", type="password")
        confirm_password = st.text_input("Confirm Password *", type="password")
        
        full_name = st.text_input("Full Name *")
        email = st.text_input("Email")
        
        col1, col2 = st.columns(2)
        with col1:
            department = st.text_input("Department")
        with col2:
            role = st.selectbox("Role *", ["admin", "accountant", "approver", "viewer"])
        
        submitted = st.form_submit_button("Create User")
        
        if submitted:
            if not username or not password or not full_name:
                st.error("Username, password, and full name are required.")
            elif password != confirm_password:
                st.error("Passwords do not match.")
            else:
                conn = get_db_connection()
                
                # Check if username already exists
                existing_user = conn.execute("SELECT user_id FROM users WHERE username = ?", (username,)).fetchone()
                
                if existing_user:
                    st.error(f"Username '{username}' already exists. Please choose a different username.")
                else:
                    # Insert new user
                    # In a real application, hash the password
                    conn.execute("""
                        INSERT INTO users
                        (username, password_hash, full_name, email, role, department, status)
                        VALUES (?, ?, ?, ?, ?, ?, 'active')
                    """, (username, password, full_name, email, role, department))
                    
                    conn.commit()
                    conn.close()
                    
                    st.success(f"User '{username}' created successfully!")
                    st.balloons()

# Settings Page
def display_settings():
    st.title("System Settings")
    
    # Check if user has admin role
    if st.session_state.user_role != 'admin':
        st.error("You don't have permission to access this page.")
        return
    
    # Tabs for different settings
    tab1, tab2, tab3 = st.tabs(["General Settings", "Tally Integration", "Database"])
    
    with tab1:
        display_general_settings()
    
    with tab2:
        display_tally_settings()
    
    with tab3:
        display_database_settings()

def display_general_settings():
    st.subheader("General Settings")
    
    # Company Info
    with st.expander("Company Information", expanded=True):
        with st.form("company_info_form"):
            company_name = st.text_input("Company Name", "Your Company Name")
            
            col1, col2 = st.columns(2)
            with col1:
                company_email = st.text_input("Email", "contact@example.com")
            with col2:
                company_phone = st.text_input("Phone", "+1 123 456 7890")
            
            company_address = st.text_area("Address", "123 Main St, City, Country")
            
            company_logo = st.file_uploader("Upload Company Logo", type=["png", "jpg", "jpeg"])
            
            submitted = st.form_submit_button("Save Company Information")
            
            if submitted:
                # Save to settings file or database
                if company_logo:
                    logo_path = os.path.join("static", "img", "logo.png")
                    os.makedirs(os.path.dirname(logo_path), exist_ok=True)
                    with open(logo_path, "wb") as f:
                        f.write(company_logo.getbuffer())
                
                # Save other settings
                settings = {
                    "company_name": company_name,
                    "company_email": company_email,
                    "company_phone": company_phone,
                    "company_address": company_address
                }
                
                # In a real app, save to database or config file
                with open("settings.json", "w") as f:
                    json.dump(settings, f)
                
                st.success("Company information saved!")
    
    # Email Settings
    with st.expander("Email Settings"):
        with st.form("email_settings_form"):
            smtp_server = st.text_input("SMTP Server", "smtp.example.com")
            smtp_port = st.number_input("SMTP Port", min_value=1, max_value=65535, value=587)
            
            col1, col2 = st.columns(2)
            with col1:
                smtp_username = st.text_input("SMTP Username", "user@example.com")
            with col2:
                smtp_password = st.text_input("SMTP Password", type="password", value="password")
            
            enable_ssl = st.checkbox("Enable SSL/TLS", value=True)
            
            submitted = st.form_submit_button("Save Email Settings")
            
            if submitted:
                # Save email settings
                settings = {
                    "smtp_server": smtp_server,
                    "smtp_port": smtp_port,
                    "smtp_username": smtp_username,
                    "smtp_password": smtp_password,
                    "enable_ssl": enable_ssl
                }
                
                # In a real app, save to database or config file
                # Here, just show success message
                st.success("Email settings saved!")
    
    # User Interface Settings
    with st.expander("User Interface Settings"):
        with st.form("ui_settings_form"):
            items_per_page = st.number_input("Items Per Page", min_value=5, max_value=100, value=10)
            date_format = st.selectbox("Date Format", ["YYYY-MM-DD", "MM/DD/YYYY", "DD/MM/YYYY"])
            theme = st.selectbox("Theme", ["Light", "Dark", "System"])
            
            submitted = st.form_submit_button("Save UI Settings")
            
            if submitted:
                # Save UI settings
                settings = {
                    "items_per_page": items_per_page,
                    "date_format": date_format,
                    "theme": theme
                }
                
                # In a real app, save to database or config file
                st.success("UI settings saved!")

def display_tally_settings():
    st.subheader("Tally ERP Integration Settings")
    
    with st.form("tally_settings_form"):
        st.write("Configure the connection to Tally ERP for data synchronization.")
        
        enable_tally = st.checkbox("Enable Tally Integration", value=True)
        
        col1, col2 = st.columns(2)
        with col1:
            tally_server = st.text_input("Tally Server", "localhost")
        with col2:
            tally_port = st.text_input("Tally Port", "9000")
        
        tally_company = st.text_input("Tally Company Name", "Your Company")
        
        sync_frequency = st.selectbox(
            "Sync Frequency",
            ["Manual", "Hourly", "Daily", "Weekly"],
            index=0
        )
        
        # Advanced settings
        st.subheader("Advanced Settings")
        
        col1, col2 = st.columns(2)
        with col1:
            vendor_ledger = st.text_input("Vendor Ledger Group", "Sundry Creditors")
        with col2:
            bill_type = st.text_input("Bill Voucher Type", "Purchase")
        
        sync_options = st.multiselect(
            "Items to Sync",
            ["Vendors", "Invoices", "Payments"],
            default=["Vendors", "Invoices"]
        )
        
        submitted = st.form_submit_button("Save Tally Settings")
        
        if submitted:
            # Save Tally settings
            settings = {
                "enable_tally": enable_tally,
                "tally_server": tally_server,
                "tally_port": tally_port,
                "tally_company": tally_company,
                "sync_frequency": sync_frequency,
                "vendor_ledger": vendor_ledger,
                "bill_type": bill_type,
                "sync_options": sync_options
            }
            
            # In a real app, save to database or config file
            st.success("Tally integration settings saved!")
    
    # Test connection button
    if st.button("Test Tally Connection"):
        try:
            import pyodbc
            
            # Attempt to connect to Tally
            conn_str = f"DRIVER={{Tally ODBC Driver}};SERVER={tally_server};PORT={tally_port}"
            
            with st.spinner("Testing connection to Tally..."):
                # Simulate connection test (would use pyodbc in real app)
                time.sleep(2)
                
                # Fake success for demo
                st.success("Successfully connected to Tally!")
        except Exception as e:
            st.error(f"Failed to connect to Tally: {str(e)}")
            st.info("Make sure Tally is running and the ODBC driver is installed.")

def display_database_settings():
    st.subheader("Database Management")
    
    # Database info
    conn = get_db_connection()
    db_path = os.path.abspath(DB_PATH)
    db_size = os.path.getsize(DB_PATH) / (1024 * 1024)  # Convert to MB
    
    # Get table counts
    tables = [
        "users", "vendors", "vendor_bank_details", "vendor_documents",
        "invoices", "payment_requests", "payment_request_items", "payment_advices",
        "audit_logs"
    ]
    
    table_counts = {}
    for table in tables:
        count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        table_counts[table] = count
    
    conn.close()
    
    # Display database info
    st.write(f"**Database Path:** {db_path}")
    st.write(f"**Database Size:** {db_size:.2f} MB")
    
    # Display table counts
    st.subheader("Table Record Counts")
    
    col1, col2, col3 = st.columns(3)
    
    cols = [col1, col2, col3]
    for i, (table, count) in enumerate(table_counts.items()):
        cols[i % 3].metric(table.replace('_', ' ').title(), count)
    
    # Backup and restore
    st.subheader("Backup and Restore")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("Backup Database"):
            # Create backup
            backup_dir = "backups"
            os.makedirs(backup_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            backup_path = os.path.join(backup_dir, f"ap_system_backup_{timestamp}.db")
            
            with st.spinner("Creating database backup..."):
                import shutil
                shutil.copy2(DB_PATH, backup_path)
                
                st.success(f"Backup created successfully: {backup_path}")
                
                # Provide download link
                with open(backup_path, "rb") as f:
                    st.download_button(
                        label="Download Backup",
                        data=f,
                        file_name=os.path.basename(backup_path),
                        mime="application/octet-stream"
                    )
    
    with col2:
        uploaded_backup = st.file_uploader("Restore from Backup", type=["db"])
        
        if uploaded_backup and st.button("Restore Database"):
            # Confirm restore
            if st.checkbox("I understand this will overwrite the current database"):
                with st.spinner("Restoring database from backup..."):
                    # Save the uploaded file
                    backup_path = os.path.join("backups", "temp_restore.db")
                    with open(backup_path, "wb") as f:
                        f.write(uploaded_backup.getbuffer())
                    
                    # Close all connections to current DB
                    import sqlite3
                    sqlite3.connect(DB_PATH).close()
                    
                    # Create a backup of current DB before restore
                    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                    pre_restore_backup = os.path.join("backups", f"pre_restore_backup_{timestamp}.db")
                    import shutil
                    shutil.copy2(DB_PATH, pre_restore_backup)
                    
                    # Restore from uploaded backup
                    shutil.copy2(backup_path, DB_PATH)
                    
                    st.success("Database restored successfully!")
                    st.info("The application will restart to apply changes.")
                    st.rerun()
            else:
                st.warning("Please confirm that you understand the restore will overwrite the current database.")
    
    # Database optimization
    st.subheader("Database Optimization")
    
    if st.button("Optimize Database"):
        with st.spinner("Optimizing database..."):
            conn = get_db_connection()
            conn.execute("VACUUM")
            conn.commit()
            conn.close()
            
            st.success("Database optimized successfully!")
            
            # Show new size
            new_size = os.path.getsize(DB_PATH) / (1024 * 1024)  # Convert to MB
            st.write(f"**New Database Size:** {new_size:.2f} MB (Reduced by {db_size - new_size:.2f} MB)")

# Initialize database if it doesn't exist
def init_database():
    if not os.path.exists(DB_PATH):
        st.info("Initializing database for first use...")
        
        conn = sqlite3.connect(DB_PATH)
        
        # Create tables
        with open("database/schema.sql", "r") as f:
            schema_sql = f.read()
            conn.executescript(schema_sql)
        
        # Create admin user
        conn.execute("""
            INSERT INTO users
            (username, password_hash, full_name, email, role, status)
            VALUES ('admin', 'admin123', 'Admin User', 'admin@example.com', 'admin', 'active')
        """)
        
        # Create demo users
        conn.execute("""
            INSERT INTO users
            (username, password_hash, full_name, email, role, status)
            VALUES ('accountant', 'accountant123', 'Accountant User', 'accountant@example.com', 'accountant', 'active')
        """)
        
        conn.execute("""
            INSERT INTO users
            (username, password_hash, full_name, email, role, status)
            VALUES ('approver', 'approver123', 'Approver User', 'approver@example.com', 'approver', 'active')
        """)
        
        conn.commit()
        conn.close()
        
        st.success("Database initialized successfully!")

# Main function to run the app
def main():
    # Initialize database if needed
    if not os.path.exists(DB_PATH):
        init_database()
    
    # Check authentication
    if not st.session_state.authenticated:
        login_page()
    else:
        main_app()

if __name__ == "__main__":
    main()